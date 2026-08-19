"""Rebuild the active collection seed from a reconciled workbook.

Stable master IDs are preserved, existing grading payloads stay attached to the
same IDs, and removed reconciliation rows are archived instead of discarded.
"""

from __future__ import annotations

import argparse
import json
import subprocess
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
CANONICAL = ROOT / "seed" / "collection.json"
APP_DATA = ROOT / "apps" / "web" / "src" / "data" / "collection.json"
RETIRED = ROOT / "seed" / "retired-v25.json"
APP_RETIRED = ROOT / "apps" / "web" / "src" / "data" / "retired-v25.json"


def cleaned(value: object) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def number(value: object) -> float | None:
    value = cleaned(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    parsed = float(value)
    return int(parsed) if parsed.is_integer() else parsed


def integer(value: object) -> int | None:
    value = cleaned(value)
    return None if value is None else int(value)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--baseline-git-ref", help="Read the prior seed from this Git ref when rebuilding reconciliation archives")
    args = parser.parse_args()

    if args.baseline_git_ref:
        previous = json.loads(subprocess.check_output(
            ["git", "show", f"{args.baseline_git_ref}:seed/collection.json"], cwd=ROOT, text=True, encoding="utf-8"
        ))
    else:
        previous = json.loads(CANONICAL.read_text(encoding="utf-8"))
    previous_by_id = {int(card["masterId"]): card for card in previous["cards"]}
    if RETIRED.exists() and not args.baseline_git_ref:
        existing_retired = json.loads(RETIRED.read_text(encoding="utf-8"))
        for record in existing_retired.get("records", []):
            card = record.get("card")
            if card:
                previous_by_id.setdefault(int(card["masterId"]), card)
    workbook = load_workbook(args.workbook, read_only=True, data_only=True)

    rows = workbook["All Physical Cards"].iter_rows(values_only=True)
    headers = {str(value).strip(): index for index, value in enumerate(next(rows)) if value is not None}
    required = {
        "ID", "Section", "Card Name", "Card No.", "Set", "Year", "Variant / Printing", "Copy",
        "Clean Raw Low", "Clean Raw High", "Clean Raw Mid", "Pile", "Sleeve", "Toploader",
        "Condition", "As-Is Low", "As-Is High", "As-Is Mid", "Notes", "Source Chat Index",
        "Source Basis", "Research URL",
    }
    missing = required - headers.keys()
    if missing:
        raise SystemExit(f"Workbook columns missing: {sorted(missing)}")

    def cell(row: tuple[object, ...], name: str) -> object:
        index = headers[name]
        return row[index] if index < len(row) else None

    active: list[dict[str, Any]] = []
    seen: set[int] = set()
    for row in rows:
        raw_id = cell(row, "ID")
        if raw_id is None:
            continue
        master_id = int(raw_id)
        if master_id in seen:
            raise SystemExit(f"Duplicate active master ID {master_id}")
        seen.add(master_id)
        old = previous_by_id.get(master_id, {})
        grading = old.get("grading")
        card = {
            "masterId": master_id,
            "category": cleaned(cell(row, "Section")),
            "name": cleaned(cell(row, "Card Name")),
            "cardNumber": cleaned(cell(row, "Card No.")),
            "setName": cleaned(cell(row, "Set")),
            "year": integer(cell(row, "Year")),
            "variant": cleaned(cell(row, "Variant / Printing")),
            "copyLabel": cleaned(cell(row, "Copy")),
            "rawLow": number(cell(row, "Clean Raw Low")),
            "rawHigh": number(cell(row, "Clean Raw High")),
            "rawMid": number(cell(row, "Clean Raw Mid")),
            "valueBucket": cleaned(cell(row, "Pile")),
            "sleeve": cleaned(cell(row, "Sleeve")),
            "toploader": cleaned(cell(row, "Toploader")),
            "logged": old.get("logged"),
            "condition": cleaned(cell(row, "Condition")),
            "asIsLow": number(cell(row, "As-Is Low")),
            "asIsHigh": number(cell(row, "As-Is High")),
            "asIsMid": number(cell(row, "As-Is Mid")),
            "notes": cleaned(cell(row, "Notes")),
            "sourceIndex": cleaned(cell(row, "Source Chat Index")),
            "source": cleaned(cell(row, "Source Basis")),
            "sourceUrl": cleaned(cell(row, "Research URL")),
            "grading": grading,
            "gradingStatus": old.get("gradingStatus", "ungraded") if grading else old.get("gradingStatus", "ungraded"),
        }
        active.append(card)

    reconciliation_rows = workbook["Duplicate Reconciliation"].iter_rows(values_only=True)
    reconciliation_headers = {
        str(value).strip(): index for index, value in enumerate(next(reconciliation_rows)) if value is not None
    }
    reasons: dict[int, dict[str, Any]] = {}
    for row in reconciliation_rows:
        index = reconciliation_headers["Master ID"]
        if index >= len(row) or row[index] is None:
            continue
        master_id = int(row[index])
        reasons[master_id] = {
            key: cleaned(row[position] if position < len(row) else None)
            for key, position in reconciliation_headers.items()
        }

    removed_ids = sorted(previous_by_id.keys() - seen)
    retired = [
        {"card": previous_by_id[master_id], "reconciliation": reasons.get(master_id)}
        for master_id in removed_ids
    ]
    unexplained = [master_id for master_id in removed_ids if not reasons.get(master_id)]
    if unexplained:
        raise SystemExit(f"Removed IDs lack reconciliation records: {unexplained}")

    summary = {
        "cardCount": len(active),
        "gradedCount": sum(1 for card in active if card.get("grading")),
        "toploaderCount": sum(1 for card in active if str(card.get("toploader") or "").lower() == "yes"),
        "rawMidTotal": round(sum(float(card.get("rawMid") or 0) for card in active), 2),
        "asIsMidTotal": round(sum(float(card.get("asIsMid") or 0) for card in active), 2),
    }
    payload = {"summary": summary, "cards": active}
    encoded = f"{json.dumps(payload, ensure_ascii=True, indent=2)}\n"
    CANONICAL.write_text(encoded, encoding="utf-8")
    APP_DATA.write_text(encoded, encoding="utf-8")
    retired_encoded = f"{json.dumps({'source': args.workbook.name, 'retiredCount': len(retired), 'records': retired}, ensure_ascii=True, indent=2)}\n"
    RETIRED.write_text(retired_encoded, encoding="utf-8")
    APP_RETIRED.write_text(retired_encoded, encoding="utf-8")
    print(
        f"wrote {summary['cardCount']} active cards, {summary['gradedCount']} active grading records, "
        f"and archived {len(retired)} reconciled rows"
    )


if __name__ == "__main__":
    main()
