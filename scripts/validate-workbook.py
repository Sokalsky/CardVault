"""Validate that the canonical JSON import still matches the source workbook."""

from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "seed" / "source" / "collection-v24.xlsx"
COLLECTION = ROOT / "seed" / "collection.json"


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def money(value: object) -> float:
    return 0.0 if value is None else float(value)


def main() -> None:
    collection = json.loads(COLLECTION.read_text(encoding="utf-8"))
    cards = {int(card["masterId"]): card for card in collection["cards"]}

    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    sheet = workbook["All Physical Cards"]
    rows = sheet.iter_rows(values_only=True)
    headers = {text(value): index for index, value in enumerate(next(rows))}
    required = {
        "ID", "Card Name", "Card No.", "Set", "Year", "Variant / Printing",
        "Copy", "Clean Raw Low", "Clean Raw High", "Clean Raw Mid",
    }
    missing = required - headers.keys()
    if missing:
        raise AssertionError(f"Workbook columns missing: {sorted(missing)}")

    workbook_ids: set[int] = set()
    for row in rows:
        raw_id = row[headers["ID"]]
        if raw_id is None:
            continue
        master_id = int(raw_id)
        if master_id in workbook_ids:
            raise AssertionError(f"Duplicate workbook ID {master_id}")
        workbook_ids.add(master_id)
        card = cards.get(master_id)
        if card is None:
            raise AssertionError(f"Workbook ID {master_id} is absent from collection.json")

        comparisons = {
            "name": "Card Name",
            "cardNumber": "Card No.",
            "setName": "Set",
            "variant": "Variant / Printing",
            "copyLabel": "Copy",
        }
        for json_key, column in comparisons.items():
            if text(card.get(json_key)) != text(row[headers[column]]):
                raise AssertionError(f"ID {master_id}: {json_key} differs from workbook")

        if card.get("year") is not None and int(card["year"]) != int(row[headers["Year"]]):
            raise AssertionError(f"ID {master_id}: year differs from workbook")
        for json_key, column in (("rawLow", "Clean Raw Low"), ("rawHigh", "Clean Raw High"), ("rawMid", "Clean Raw Mid")):
            if not math.isclose(money(card[json_key]), money(row[headers[column]]), abs_tol=0.0001):
                raise AssertionError(f"ID {master_id}: {json_key} differs from workbook")

    if len(workbook_ids) != 860 or workbook_ids != set(cards):
        raise AssertionError(f"Expected the same 860 IDs; workbook={len(workbook_ids)} json={len(cards)}")
    print("validated source workbook: 860 distinct physical cards match canonical JSON identities and raw prices")


if __name__ == "__main__":
    main()
